import openpyxl

excel_path = "/Users/hoangvan/Downloads/BẢN ĐỒ SETUP_TỔNG HỢP.xlsx"
wb = openpyxl.load_workbook(excel_path, read_only=True)

for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    print(f"\n--- Sheet: {sheetname} ---")
    # Get first 3 rows
    rows = list(sheet.iter_rows(max_row=3, values_only=True))
    for idx, r in enumerate(rows):
        print(f"Row {idx}: {r[:10]}")
