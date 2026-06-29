import json
import os
import openpyxl
import re

excel_path = "/Users/hoangvan/Downloads/BẢN ĐỒ SETUP_TỔNG HỢP.xlsx"
json_path = "/Users/hoangvan/.gemini/antigravity/scratch/autoshop-setup-map/src/data/shops.json"
artifact_path = "/Users/hoangvan/.gemini/antigravity/brain/c68d7c20-27da-4738-bf75-cbe5f4dcd882/excel_comparison_results.md"

def normalize_name(name):
    if not name:
        return ""
    name = name.lower()
    for word in ["cafe", "coffee", "cf", "tea", "&", "trà sữa", "kem trứng", "lilli", "setup", "set up", "kafe", "caffe"]:
        name = name.replace(word, " ")
    # Keep only alphanumeric
    return "".join(c for c in name if c.isalnum())

def normalize_address(addr):
    if not addr:
        return ""
    addr = addr.lower()
    return "".join(c for c in addr if c.isalnum())

# Load existing JSON data
with open(json_path, 'r', encoding='utf-8') as f:
    json_shops = json.load(f)

# Build index of existing shops
existing_by_norm_name = {}
for s in json_shops:
    name = s.get('name', '')
    norm = normalize_name(name)
    if norm:
        if norm not in existing_by_norm_name:
            existing_by_norm_name[norm] = []
        existing_by_norm_name[norm].append(s)

wb = openpyxl.load_workbook(excel_path, data_only=True)

excel_shops = []
for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        continue
    
    header = rows[0]
    name_idx = -1
    addr_idx = -1
    for col_idx, col in enumerate(header):
        if col:
            col_str = str(col).strip().upper()
            if 'TÊN QUÁN' in col_str or 'TEN QUAN' in col_str:
                name_idx = col_idx
            elif 'ĐỊA CHỈ' in col_str or 'DIA CHI' in col_str:
                addr_idx = col_idx
                
    if name_idx == -1 or addr_idx == -1:
        continue
        
    for r_idx, row in enumerate(rows[1:], start=2):
        if len(row) > max(name_idx, addr_idx):
            name = row[name_idx]
            addr = row[addr_idx]
            if name and str(name).strip() and addr and str(addr).strip():
                name_str = str(name).strip()
                addr_str = str(addr).strip()
                # Skip placeholder names or entries
                if name_str.lower() in ['x', 'đã đóng cửa', 'setup menu']:
                    continue
                excel_shops.append({
                    'sheet': sheetname,
                    'row': r_idx,
                    'name': name_str,
                    'address': addr_str
                })

different_address_shops = []
missing_shops = []

for es in excel_shops:
    norm_name = normalize_name(es['name'])
    if not norm_name:
        continue
        
    if norm_name in existing_by_norm_name:
        matched = False
        es_addr_norm = normalize_address(es['address'])
        
        for js in existing_by_norm_name[norm_name]:
            js_addr_norm = normalize_address(js.get('address', ''))
            
            # check similarity
            if es_addr_norm in js_addr_norm or js_addr_norm in es_addr_norm:
                matched = True
                break
                
        if not matched:
            different_address_shops.append({
                'excel': es,
                'json_options': existing_by_norm_name[norm_name]
            })
    else:
        missing_shops.append(es)

# Write to Markdown Artifact
with open(artifact_path, 'w', encoding='utf-8') as f:
    f.write("# Kết quả đối chiếu BẢN ĐỒ SETUP_TỔNG HỢP\n\n")
    f.write("Báo cáo đối chiếu dữ liệu giữa file Excel `BẢN ĐỒ SETUP_TỔNG HỢP.xlsx` trên máy và cơ sở dữ liệu `shops.json` hiện tại của website.\n\n")
    
    f.write("## Thống kê tổng quan\n")
    f.write(f"- **Số lượng quán trên website hiện tại:** {len(json_shops)} quán\n")
    f.write(f"- **Số lượng quán hợp lệ trích xuất từ Excel:** {len(excel_shops)} quán\n")
    f.write(f"- **Số lượng quán trùng tên nhưng khác địa chỉ:** {len(different_address_shops)} quán\n")
    f.write(f"- **Số lượng quán hoàn toàn chưa có trên website:** {len(missing_shops)} quán\n\n")
    
    f.write("## 1. Quán trùng tên nhưng khác địa chỉ (Cần rà soát cập nhật)\n")
    f.write("Dưới đây là các quán đã có trên website (trùng tên) nhưng địa chỉ ghi nhận trong file Excel khác biệt đáng kể hoặc chi tiết hơn.\n\n")
    
    if different_address_shops:
        f.write("| Tên quán | Địa chỉ trên Web | Địa chỉ trên Excel (Sheet, Dòng) |\n")
        f.write("| --- | --- | --- |\n")
        for item in different_address_shops:
            es = item['excel']
            jo = item['json_options'][0]
            f.write(f"| **{es['name']}** | {jo.get('address', 'N/A')} | {es['address']} (*{es['sheet']}*, Dòng {es['row']}) |\n")
    else:
        f.write("*Không có quán nào trùng tên khác địa chỉ.*\n")
        
    f.write("\n## 2. Quán hoàn toàn chưa có trên website (Cần thêm mới)\n")
    f.write("Dưới đây là danh sách các quán được tìm thấy trong file Excel nhưng không trùng tên với bất kỳ quán nào trên website.\n\n")
    
    if missing_shops:
        f.write("| STT | Tên quán | Địa chỉ | Vị trí Excel (Sheet, Dòng) |\n")
        f.write("| --- | --- | --- | --- |\n")
        for idx, es in enumerate(missing_shops, 1):
            f.write(f"| {idx} | **{es['name']}** | {es['address']} | {es['sheet']}, Dòng {es['row']} |\n")
    else:
        f.write("*Không có quán mới nào cần thêm.*\n")

print(f"Report written to {artifact_path}")
