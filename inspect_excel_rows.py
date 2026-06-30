import openpyxl

file_path = "/Users/hoangvan/Downloads/BẢN ĐỒ SETUP_TỔNG HỢP.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=True)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n--- Sheet: {sheet_name} ---")
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    print(f"Headers: {headers}")
    
    # Check first 5 rows
    row_count = 0
    for row in sheet.iter_rows(min_row=2, max_row=10):
        row_vals = [cell.value for cell in row]
        if any(row_vals):
            print(f"Row {row_count+2}: {row_vals[:12]}")
            row_count += 1
        if row_count >= 3:
            break
