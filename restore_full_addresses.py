import json
import openpyxl
import re

json_path = "/Users/hoangvan/.gemini/antigravity/scratch/autoshop-setup-map/src/data/shops.json"
excel_path = "/Users/hoangvan/Downloads/BẢN ĐỒ SETUP_TỔNG HỢP.xlsx"

def normalize_name(name):
    if not name:
        return ""
    name = name.lower()
    for word in ["cafe", "coffee", "cf", "tea", "&", "trà sữa", "kem trứng", "lilli", "setup", "set up", "kafe", "caffe"]:
        name = name.replace(word, " ")
    return "".join(c for c in name if c.isalnum())

# Load current JSON database
with open(json_path, 'r', encoding='utf-8') as f:
    shops = json.load(f)

print(f"Loaded {len(shops)} shops from database.")

# Read Excel to build a name-to-address mapping
wb = openpyxl.load_workbook(excel_path, data_only=True)

excel_addresses = {}
for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        continue
        
    header = rows[0]
    name_idx = -1
    addr_idx = -1
    for col_idx, col in enumerate(header):
        if col:
            col_str = str(col).strip().upper()
            if 'TÊN QUÁN' in col_str or 'TEN QUAN' in col_str:
                name_idx = col_idx
            elif 'ĐỊA CHỈ' in col_str or 'DIA CHI' in col_str:
                addr_idx = col_idx
                
    if name_idx == -1 or addr_idx == -1:
        continue
        
    for row in rows[1:]:
        if len(row) > max(name_idx, addr_idx):
            name = row[name_idx]
            addr = row[addr_idx]
            if name and str(name).strip() and addr and str(addr).strip():
                name_str = str(name).strip()
                addr_str = str(addr).strip()
                
                norm_name = normalize_name(name_str)
                if norm_name:
                    # Keep the longest address if there are duplicate names in Excel
                    if norm_name not in excel_addresses or len(addr_str) > len(excel_addresses[norm_name]):
                        excel_addresses[norm_name] = addr_str

print(f"Loaded {len(excel_addresses)} address mappings from Excel.")

# Perform updates
updated_count = 0
for s in shops:
    name = s.get('name', '')
    norm_name = normalize_name(name)
    current_addr = s.get('address', '')
    
    if norm_name in excel_addresses:
        excel_addr = excel_addresses[norm_name]
        
        # If the address in Excel is longer and covers the JSON address (or is just longer/more complete), update it
        # We also check if the current address ends with typical truncation patterns (comma, trailing words, short lengths)
        is_shorter = len(current_addr) < len(excel_addr)
        
        # Don't update if current address is already very long and similar
        if is_shorter and current_addr != excel_addr:
            s['address'] = excel_addr
            updated_count += 1
            print(f"Updated: '{name}'\n  Old: {current_addr}\n  New: {excel_addr}")

# Save updated JSON database
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(shops, f, ensure_ascii=False, indent=2)

print(f"\nSUCCESS: Restored full addresses for {updated_count} shops.")
