import sys
import os

excel_path = "/Users/hoangvan/Downloads/BẢN ĐỒ SETUP_TỔNG HỢP.xlsx"

if not os.path.exists(excel_path):
    print(f"Error: File not found at {excel_path}")
    sys.exit(1)

print("File exists. Trying to read...")

try:
    import openpyxl
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    print("Sheets:", wb.sheetnames)
    sheet = wb.active
    print("Active sheet:", sheet.title)
    
    # Read first 10 rows
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if r_idx < 15:
            print(f"Row {r_idx}: {row}")
        else:
            break
except Exception as e:
    print("Error reading with openpyxl:", e)
    
    try:
        import pandas as pd
        df = pd.read_excel(excel_path)
        print("Pandas read success. Shape:", df.shape)
        print("Columns:", df.columns.tolist())
        print(df.head(10))
    except Exception as e2:
        print("Error reading with pandas:", e2)
