import openpyxl

file_path = "/Users/hoangvan/Downloads/BẢN ĐỒ SETUP_TỔNG HỢP.xlsx"
try:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    for name in wb.sheetnames:
        sheet = wb[name]
        print(f"Sheet: {name}")
        # Print first row
        row1 = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        print(f"Columns: {row1}")
except Exception as e:
    print(f"Error reading Excel: {e}")
