import json
import os
import openpyxl
import random

excel_path = "/Users/hoangvan/Downloads/BẢN ĐỒ SETUP_TỔNG HỢP.xlsx"
json_path = "/Users/hoangvan/.gemini/antigravity/scratch/autoshop-setup-map/src/data/shops.json"

PROVINCE_COORDINATES = {
    "AN GIANG": (10.5, 105.2),
    "BÀ RỊA - VŨNG TÀU": (10.5, 107.2),
    "VŨNG TÀU": (10.4, 107.1),
    "BẠC LIÊU": (9.3, 105.5),
    "BẮC GIANG": (21.3, 106.2),
    "BẮC KẠN": (22.3, 105.8),
    "BẮC NINH": (21.2, 106.1),
    "BẾN TRE": (10.2, 106.4),
    "BÌNH DƯƠNG": (11.1, 106.7),
    "BÌNH ĐỊNH": (13.8, 109.1),
    "BÌNH PHƯỚC": (11.7, 106.9),
    "BÌNH THUẬN": (11.1, 108.1),
    "CÀ MAU": (9.2, 105.1),
    "CẦN THƠ": (10.0, 105.7),
    "CAO BẰNG": (22.7, 106.3),
    "ĐÀ NẴNG": (16.0, 108.2),
    "ĐẮK LẮK": (12.7, 108.1),
    "ĐẮK NÔNG": (12.1, 107.7),
    "ĐIỆN BIÊN": (21.8, 103.0),
    "ĐỒNG NAI": (11.0, 107.0),
    "ĐỒNG THÁP": (10.5, 105.6),
    "GIA LAI": (14.0, 108.0),
    "HÀ GIANG": (22.8, 104.9),
    "HÀ NAM": (20.5, 105.9),
    "HÀ NỘI": (21.0, 105.8),
    "HÀ TĨNH": (18.3, 105.9),
    "HẢI DƯƠNG": (20.9, 106.3),
    "HẢI PHÒNG": (20.8, 106.7),
    "HẬU GIANG": (9.8, 105.6),
    "HÒA BÌNH": (20.7, 105.3),
    "HƯNG YÊN": (20.6, 106.1),
    "KHÁNH HÒA": (12.3, 109.1),
    "KIÊN GIANG": (10.0, 105.1),
    "KON TUM": (14.3, 108.0),
    "LAI CHÂU": (22.3, 103.0),
    "LẠNG SƠN": (21.8, 106.7),
    "LÀO CAI": (22.4, 103.9),
    "LÂM ĐỒNG": (11.6, 107.8),
    "ĐÀ LẠT": (11.9, 108.4),
    "LONG AN": (10.7, 106.3),
    "NAM ĐỊNH": (20.4, 106.2),
    "NGHỆ AN": (19.2, 105.1),
    "NINH BÌNH": (20.2, 105.9),
    "NINH THUẬN": (11.7, 108.9),
    "PHÚ THỌ": (21.3, 105.2),
    "PHÚ YÊN": (13.1, 109.1),
    "QUẢNG BÌNH": (17.5, 106.3),
    "QUẢNG NAM": (15.7, 107.9),
    "QUẢNG NGÃI": (15.1, 108.8),
    "QUẢNG NINH": (21.0, 107.3),
    "QUẢNG TRỊ": (16.7, 107.1),
    "SÓC TRĂNG": (9.6, 106.0),
    "SƠN LA": (21.2, 103.7),
    "TÂY NINH": (11.4, 106.1),
    "THÁI BÌNH": (20.5, 106.4),
    "THÁI NGUYÊN": (21.6, 105.8),
    "THANH HÓA": (19.8, 105.8),
    "THỪA THIÊN HUẾ": (16.5, 107.5),
    "HUẾ": (16.5, 107.5),
    "TIỀN GIANG": (10.4, 106.2),
    "HỒ CHÍ MINH": (10.8, 106.6),
    "TP HCM": (10.8, 106.6),
    "TRÀ VINH": (9.9, 106.3),
    "TUYÊN QUANG": (21.8, 105.2),
    "VĨNH LONG": (10.2, 105.9),
    "VĨNH PHÚC": (21.3, 105.5),
    "YÊN BÁI": (21.7, 104.9)
}

# Helper to normalize province name
def clean_province(p):
    if not p:
        return ""
    p = str(p).strip().upper()
    p = p.replace("TP.", "").replace("TỈNH", "").strip()
    p = re.sub(r'\s+', ' ', p)
    if p in ["HÀ NỘI", "HN"]:
        return "HÀ NỘI"
    if p in ["TP HCM", "TP.HCM", "HỒ CHÍ MINH", "HCM", "SÀI GÒN"]:
        return "HỒ CHÍ MINH"
    if p in ["HẢI PHÒNG", "HP"]:
        return "HẢI PHÒNG"
    if p in ["ĐÀ NẴNG", "ĐN"]:
        return "ĐÀ NẴNG"
    if p in ["BÀ RỊA VŨNG TÀU", "VŨNG TÀU", "BRVT", "BR-VT"]:
        return "BÀ RỊA - VŨNG TÀU"
    return p

import re

def detect_province_from_address(addr):
    if not addr:
        return "HÀ NỘI"
    addr_upper = addr.upper()
    for prov in PROVINCE_COORDINATES.keys():
        if prov in addr_upper:
            return prov
    return "HÀ NỘI" # default if not found

def normalize_name(name):
    if not name:
        return ""
    name = name.lower()
    for word in ["cafe", "coffee", "cf", "tea", "&", "trà sữa", "kem trứng", "lilli", "setup", "set up", "kafe", "caffe"]:
        name = name.replace(word, " ")
    return "".join(c for c in name if c.isalnum())

def normalize_address(addr):
    if not addr:
        return ""
    return "".join(c for c in addr.lower() if c.isalnum())

# Load existing JSON
with open(json_path, 'r', encoding='utf-8') as f:
    json_shops = json.load(f)

# Find max current numeric ID
max_num = 0
for shop in json_shops:
    sid = shop.get('id', '')
    match = re.search(r'MI-(\d+)', sid)
    if match:
        num = int(match.group(1))
        if num > max_num:
            max_num = num

print(f"Max ID in JSON is: MI-{max_num:03d}")

# Index existing names
existing_norm_names = set(normalize_name(s.get('name', '')) for s in json_shops if s.get('name'))

wb = openpyxl.load_workbook(excel_path, data_only=True)

new_shops_added = 0
for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        continue
    
    header = rows[0]
    name_idx = -1
    addr_idx = -1
    prov_idx = -1
    invest_idx = -1
    model_idx = -1
    cust_idx = -1
    
    for col_idx, col in enumerate(header):
        if col:
            col_str = str(col).strip().upper()
            if 'TÊN QUÁN' in col_str or 'TEN QUAN' in col_str:
                name_idx = col_idx
            elif 'ĐỊA CHỈ' in col_str or 'DIA CHI' in col_str:
                addr_idx = col_idx
            elif 'TỈNH THÀNH' in col_str or 'TINH THANH' in col_str:
                prov_idx = col_idx
            elif 'TỔNG ĐẦU TƯ' in col_str or 'TONG DAU TU' in col_str:
                invest_idx = col_idx
            elif 'MÔ HÌNH' in col_str or 'MO HINH' in col_str:
                model_idx = col_idx
            elif 'KHÁCH HÀNG' in col_str or 'KHACH HANG' in col_str:
                cust_idx = col_idx
                
    if name_idx == -1 or addr_idx == -1:
        continue
        
    # Map sheet to region
    sheet_upper = sheetname.upper()
    if 'BẮC' in sheet_upper:
        region = 'Miền Bắc'
    elif 'TRUNG' in sheet_upper:
        region = 'Miền Trung'
    elif 'NAM' in sheet_upper:
        region = 'Miền Nam'
    else:
        region = 'Miền Bắc' # default fallback
        
    for r_idx, row in enumerate(rows[1:], start=2):
        if len(row) > max(name_idx, addr_idx):
            name = row[name_idx]
            addr = row[addr_idx]
            if name and str(name).strip() and addr and str(addr).strip():
                name_str = str(name).strip()
                addr_str = str(addr).strip()
                
                # Skip placeholder names
                if name_str.lower() in ['x', 'đã đóng cửa', 'setup menu']:
                    continue
                    
                norm_name = normalize_name(name_str)
                if not norm_name:
                    continue
                    
                if norm_name not in existing_norm_names:
                    # Determine province
                    prov_str = ""
                    if prov_idx != -1 and len(row) > prov_idx and row[prov_idx]:
                        prov_str = clean_province(row[prov_idx])
                    if not prov_str:
                        prov_str = clean_province(detect_province_from_address(addr_str))
                        
                    # Handle province mapping anomalies
                    if prov_str == "TP HCM" or prov_str == "HCM" or prov_str == "SÀI GÒN":
                        prov_str = "HỒ CHÍ MINH"
                        
                    # Get coordinates
                    base_coords = PROVINCE_COORDINATES.get(prov_str, (16.0, 106.5))
                    lat = base_coords[0] + random.uniform(-0.02, 0.02)
                    lng = base_coords[1] + random.uniform(-0.02, 0.02)
                    
                    # Optional columns
                    invest_str = str(row[invest_idx]).strip() if (invest_idx != -1 and len(row) > invest_idx and row[invest_idx]) else ""
                    model_str = str(row[model_idx]).strip() if (model_idx != -1 and len(row) > model_idx and row[model_idx]) else ""
                    cust_str = str(row[cust_idx]).strip() if (cust_idx != -1 and len(row) > cust_idx and row[cust_idx]) else ""
                    
                    max_num += 1
                    new_id = f"MI-{max_num:03d}"
                    
                    new_shop = {
                        "id": new_id,
                        "region": region,
                        "province": prov_str,
                        "name": name_str,
                        "address": addr_str,
                        "investment": invest_str,
                        "model": model_str,
                        "target_customers": cust_str,
                        "images": [],
                        "info_card_url": "",
                        "lat": lat,
                        "lng": lng
                    }
                    
                    json_shops.append(new_shop)
                    existing_norm_names.add(norm_name)
                    new_shops_added += 1

# Save updated database
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(json_shops, f, ensure_ascii=False, indent=2)

print(f"Successfully added {new_shops_added} new shops to {json_path}!")
print(f"Total shops now in database: {len(json_shops)}")
